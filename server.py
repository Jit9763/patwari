import http.server
import socketserver
import json
import os
import openpyxl
from datetime import datetime
import random
import urllib.request

GOOGLE_SCRIPT_URL = 'https://script.google.com/macros/s/AKfycbyc226bwat0bvih-TWfx89He5HYC2O69NivHdt3KgUeuT_wXUrGzjNmwM8XpjABCvA1tg/exec'

def forward_to_google_sheet(action, payload_data):
    body = {
        "action": action,
        "password": "1520",
        "payload": payload_data
    }
    req_body = json.dumps(body, ensure_ascii=False).encode('utf-8')
    req = urllib.request.Request(
        GOOGLE_SCRIPT_URL,
        data=req_body,
        headers={
            'Content-Type': 'application/json; charset=utf-8',
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)'
        }
    )
    try:
        # Request with a 12 second timeout to be safe
        with urllib.request.urlopen(req, timeout=12) as response:
            res_content = response.read().decode('utf-8')
            print(f"Google Sheet Sync ({action}) Success: {res_content[:150]}...")
            return json.loads(res_content)
    except Exception as e:
        print(f"Google Sheet Sync ({action}) Failed: {e}")
        return {"status": "error", "message": str(e)}
def normalize_mobile(m_val):
    if m_val is None:
        return ""
    m_str = str(m_val).strip()
    if m_str.endswith('.0'):
        m_str = m_str[:-2]
    # Remove any non-digit chars
    m_str = "".join(c for c in m_str if c.isdigit())
    return m_str


PORT = 8000
EXCEL_PATH = 'c:/Users/jiten/Desktop/class11/patwari/patwari_reports.xlsx'

class CustomHandler(http.server.SimpleHTTPRequestHandler):
    def end_headers(self):
        # Add CORS headers just in case
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'GET, POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
        super().end_headers()

    def do_OPTIONS(self):
        self.send_response(200, "OK")
        self.end_headers()

    def do_POST(self):
        if self.path == '/api':
            content_length = int(self.headers['Content-Length'])
            post_data = self.rfile.read(content_length)
            
            try:
                payload = json.loads(post_data.decode('utf-8'))
                action = payload.get('action')
                response_data = self.handle_api_action(action, payload)
                
                self.send_response(200)
                self.send_header('Content-Type', 'application/json; charset=utf-8')
                self.end_headers()
                self.wfile.write(json.dumps(response_data, ensure_ascii=False).encode('utf-8'))
            except Exception as e:
                print("API ERROR:", e)
                self.send_response(500)
                self.send_header('Content-Type', 'application/json; charset=utf-8')
                self.end_headers()
                err_resp = {"status": "error", "message": str(e)}
                self.wfile.write(json.dumps(err_resp).encode('utf-8'))
        else:
            self.send_response(404)
            self.end_headers()

    def handle_api_action(self, action, payload):
        if not os.path.exists(EXCEL_PATH):
            raise FileNotFoundError(f"Excel file not found at {EXCEL_PATH}. Please run create_initial_excel.py first.")

        wb = openpyxl.load_workbook(EXCEL_PATH)
        
        # Helper to convert sheet rows to dictionaries
        def get_sheet_records(sheet):
            if sheet.max_row < 2:
                return []
            headers = [cell.value for cell in sheet[1]]
            records = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if any(row):  # skip empty rows
                    records.append(dict(zip(headers, row)))
            return records

        # Helper to update or append row
        def save_record(sheet, data, key_col="report_id"):
            headers = [cell.value for cell in sheet[1]]
            key_val = data.get(key_col)
            
            target_row_idx = None
            if key_val:
                for r_idx in range(2, sheet.max_row + 1):
                    cell_val = sheet.cell(row=r_idx, column=headers.index(key_col) + 1).value
                    if str(cell_val).strip() == str(key_val).strip():
                        target_row_idx = r_idx
                        break
            
            if target_row_idx is None:
                target_row_idx = sheet.max_row + 1
            
            for key, val in data.items():
                if key not in headers:
                    # Dynamically add new column to sheet headers
                    new_col_idx = sheet.max_column + 1
                    sheet.cell(row=1, column=new_col_idx, value=key)
                    headers.append(key)
                col_idx = headers.index(key) + 1
                # Handle json lists/dicts as strings
                if isinstance(val, (list, dict)):
                    val = json.dumps(val, ensure_ascii=False)
                sheet.cell(row=target_row_idx, column=col_idx, value=val)

        if action == 'getPatwariReport':
            mobile = normalize_mobile(payload.get('mobile_no'))
            if not mobile:
                return {"status": "error", "message": "मोबाईल नम्बर खाली है।"}
            
            # 1. Search in Saved Reports
            sheet2 = wb['SavedReports']
            reports = get_sheet_records(sheet2)
            found_report = next((r for r in reversed(reports) if normalize_mobile(r.get('mobile_no')) == mobile), None)
            
            if found_report:
                return {"status": "success", "source": "saved_report", "data": found_report}
            
            # 2. Search in PrefillData
            sheet1 = wb['PrefillData']
            prefill_db = get_sheet_records(sheet1)
            found_prefill = next((p for p in prefill_db if normalize_mobile(p.get('mobile_no')) == mobile), None)
            
            if found_prefill:
                return {"status": "success", "source": "prefill_db", "data": found_prefill}
            else:
                return {"status": "not_found", "message": "इस मोबाईल नम्बर के लिए कोई रिकॉर्ड नहीं मिला।"}

        elif action == 'savePatwari':
            sheet2 = wb['SavedReports']
            report = payload.get('payload', {})
            
            # Generate report_id if not present (creating new vs editing existing)
            report_id = report.get('report_id')
            if not report_id:
                report_id = 'REP-MOCK-' + datetime.now().strftime('%Y%m%d') + '-' + str(random.randint(1000, 9999))
                report['report_id'] = report_id
            
            # Always reset status to Pending and clear SDM verification details when Patwari saves
            report['status'] = "Pending"
            report['sdm_name'] = ""
            report['sdm_comments'] = ""
            report['sdm_signature_base64'] = ""
            report['sdm_timestamp'] = ""
            report['sdm_report_id'] = ""

            report['timestamp'] = datetime.now().isoformat()
            
            save_record(sheet2, report, key_col="report_id")
            wb.save(EXCEL_PATH)
            
            # Forward to Google Sheets database
            google_msg = ""
            try:
                sync_res = forward_to_google_sheet('savePatwari', report)
                if sync_res.get('status') == 'success':
                    google_msg = " और गूगल शीट में भी डेटा सुरक्षित हो गया है।"
                else:
                    google_msg = f" (गूगल शीट सिंक विफल: {sync_res.get('message')})"
            except Exception as e:
                google_msg = f" (गूगल शीट सिंक त्रुटि: {str(e)})"
            
            return {"status": "success", "message": f"रिपोर्ट लोकल एक्सेल डेटाबेस में सुरक्षित हो गई है{google_msg}", "report_id": report_id}

        elif action == 'getPendingReports':
            sheet2 = wb['SavedReports']
            reports = get_sheet_records(sheet2)
            # Filter reports that do not have an SDM signature/comment or status is not Verified
            pending = [r for r in reports if not r.get('sdm_name') or r.get('status') != "Verified"]
            return {"status": "success", "reports": pending}

        elif action == 'saveSdm':
            sheet2 = wb['SavedReports']
            sdm_payload = payload.get('payload', {})
            patwari_report_id = sdm_payload.get('patwari_report_id') or sdm_payload.get('report_id')
            
            sdm_report_id = sdm_payload.get('sdm_report_id')
            if not sdm_report_id:
                sdm_report_id = 'SDM-MOCK-' + datetime.now().strftime('%Y%m%d') + '-' + str(random.randint(1000, 9999))
                sdm_payload['sdm_report_id'] = sdm_report_id
            
            # Find the row in SavedReports sheet to update
            headers = [cell.value for cell in sheet2[1]]
            target_row_idx = None
            for r_idx in range(2, sheet2.max_row + 1):
                cell_val = sheet2.cell(row=r_idx, column=headers.index('report_id') + 1).value
                if str(cell_val).strip() == str(patwari_report_id).strip():
                    target_row_idx = r_idx
                    break
                    
            if not target_row_idx:
                return {"status": "error", "message": "सम्बन्धित पटवारी रिपोर्ट नहीं मिली।"}
                
            sdm_payload['report_id'] = patwari_report_id
            sdm_payload['sdm_timestamp'] = datetime.now().isoformat()
            sdm_payload['status'] = "Verified"
            
            # Save the entire updated record (including both modified patwari fields and SDM fields)
            save_record(sheet2, sdm_payload, key_col="report_id")
            wb.save(EXCEL_PATH)
            
            # Forward to Google Sheets database
            google_msg = ""
            try:
                sync_res = forward_to_google_sheet('saveSdm', sdm_payload)
                if sync_res.get('status') == 'success':
                    google_msg = " और गूगल शीट में भी सत्यापन स्वीकृत हो गया है।"
                else:
                    google_msg = f" (गूगल शीट सिंक विफल: {sync_res.get('message')})"
            except Exception as e:
                google_msg = f" (गूगल शीट सिंक त्रुटि: {str(e)})"
                
            return {"status": "success", "message": f"एसडीएम सत्यापन लोकल एक्सेल डेटाबेस में सुरक्षित हो गया है{google_msg}", "sdm_report_id": sdm_report_id}

        elif action == 'getAdminData':
            sheet2 = wb['SavedReports']
            reports = get_sheet_records(sheet2)
            
            total_patwari = len(reports)
            sdm_reports = []
            
            for r in reports:
                if r.get('sdm_name') and r.get('status') == "Verified":
                    # Extract SDM specific fields to match Apps Script format
                    sdm_reports.append({
                        "patwari_report_id": r.get('report_id'),
                        "sdm_report_id": r.get('sdm_report_id'),
                        "sdm_name": r.get('sdm_name'),
                        "sdm_comments": r.get('sdm_comments'),
                        "sdm_signature_base64": r.get('sdm_signature_base64'),
                        "sdm_timestamp": r.get('sdm_timestamp')
                    })
                    
            total_sdm = len(sdm_reports)
            total_pending = max(0, total_patwari - total_sdm)
            
            return {
                "status": "success",
                "summary": {
                    "total_patwari_reports": total_patwari,
                    "total_sdm_reports": total_sdm,
                    "total_pending_approvals": total_pending
                },
                "patwari_reports": reports,
                "sdm_reports": sdm_reports
            }
        else:
            return {"status": "error", "message": f"अज्ञात कार्य: {action}"}

if __name__ == '__main__':
    # Change working directory to the directory of this file
    os.chdir(os.path.dirname(os.path.abspath(__file__)))
    
    Handler = CustomHandler
    # Enable reuse address to avoid 'address already in use' errors
    socketserver.TCPServer.allow_reuse_address = True
    
    with socketserver.TCPServer(("", PORT), Handler) as httpd:
        print(f"Patwari Portal Server running at http://localhost:{PORT}/")
        print(f"Serving files from: {os.getcwd()}")
        print(f"Excel database path: {EXCEL_PATH}")
        try:
            httpd.serve_forever()
        except KeyboardInterrupt:
            print("\nStopping server...")
